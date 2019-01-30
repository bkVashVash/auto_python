# todo  正则提取到的数据， 添加自动写入excel的功能

import openpyxl
import os

'''
["{'errCode': 10000, 'errMsg': '每日商报2018年12月第20181229期文章采集成功'}",
'''
import datetime

cc = {"xxx报":"xxx期"}
'''  周日  ---  6   '''
date1 = datetime.date.weekday(datetime.datetime.now())

# todo 转到文件所在路径
os.chdir(r'C:\Users\bk201\Desktop')
#print(os.getcwd())
wb = openpyxl.load_workbook('（总）文章版1227 .xlsx')
#print(type(wb))
sheet = wb.active

for i in range(2,sheet.max_row):
    paper_name = sheet.cell(row=i,column=2).value
    if paper_name in cc:
        # 星期数从0开始
        sheet.cell(row=i,column=6+1+date1).value = cc[paper_name]
# 保存到新表中
wb.save('example.xlsx')